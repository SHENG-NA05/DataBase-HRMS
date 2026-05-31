import openpyxl

wb = openpyxl.load_workbook('人力資源派遣公司管理系統_DB設計表.xlsx')
for sheetname in wb.sheetnames:
    print(f"Sheet Name: {sheetname}")
    # Try to decode or display normally if possible
    try:
        decoded_name = sheetname.encode('utf-8').decode('utf-8')
    except Exception:
        decoded_name = sheetname
    
    sheet = wb[sheetname]
    print(f"--- {decoded_name} (rows: {sheet.max_row}, cols: {sheet.max_column}) ---")
    for r in range(1, min(sheet.max_row + 1, 15)):
        row_vals = [cell.value for cell in sheet[r]]
        if any(row_vals):
            print(row_vals)
    print("\n")
